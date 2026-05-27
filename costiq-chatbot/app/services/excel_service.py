"""
Excel export service.
Builds a multi-sheet enriched Excel workbook from all CostIQ API data.
Called when the user asks the chatbot to generate a report.
"""
import io
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Column display maps ────────────────────────────────────────────────────

FOOD_COLS = {
    "id": "ID", "entryDate": "Entry Date", "foodItem.name": "Food Item",
    "foodItem.sku": "SKU", "foodItem.category.name": "Category",
    "supplier.name": "Supplier", "country.name": "Country",
    "country.region.name": "Region", "fiscalPeriod.periodName": "Fiscal Period",
    "costCenter.code": "Cost Center", "quantity": "Quantity",
    "unitCostUsd": "Unit Cost (USD)", "totalCostUsd": "Total Cost (USD)",
    "invoiceRef": "Invoice Ref", "poNumber": "PO Number", "notes": "Notes",
}

PKG_COLS = {
    "id": "ID", "entryDate": "Entry Date", "packagingItem.name": "Packaging Item",
    "packagingItem.sku": "SKU", "packagingItem.packagingType.name": "Type",
    "packagingItem.packagingType.material": "Material",
    "supplier.name": "Supplier", "country.name": "Country",
    "country.region.name": "Region", "fiscalPeriod.periodName": "Fiscal Period",
    "quantity": "Quantity (units)", "unitCostUsd": "Unit Cost (USD)",
    "totalCostUsd": "Total Cost (USD)", "invoiceRef": "Invoice Ref",
}

TOY_COLS = {
    "id": "ID", "entryDate": "Date", "campaign.name": "Campaign",
    "toyItem.name": "Toy Item", "toyItem.toyCategory.name": "Category",
    "country.name": "Country", "country.region.name": "Region",
    "supplier.name": "Supplier", "fiscalPeriod.periodName": "Fiscal Period",
    "distributionChannel": "Channel", "quantity": "Qty",
    "unitCostUsd": "Unit Cost (USD)", "totalCostUsd": "Total Cost (USD)",
    "invoiceRef": "Invoice Ref",
}

MKT_COLS = {
    "id": "ID", "entryDate": "Date", "campaign.name": "Campaign",
    "costType": "Cost Type", "vendorName": "Vendor",
    "costCenter.code": "Cost Center", "fiscalPeriod.periodName": "Fiscal Period",
    "amountUsd": "Amount (USD)", "invoiceRef": "Invoice Ref", "description": "Notes",
}

CAMPAIGN_COLS = {
    "id": "ID", "name": "Campaign Name", "campaignCode": "Code",
    "status": "Status", "targetRegion": "Target Region",
    "startDate": "Start Date", "endDate": "End Date",
    "totalBudgetUsd": "Total Budget (USD)", "description": "Description",
}

SUPPLIER_COLS = {
    "id": "ID", "supplierCode": "Code", "name": "Supplier Name",
    "contactName": "Contact", "contactEmail": "Email",
    "country.name": "Country", "active": "Active",
}


# ── Utility: flatten nested dict fields ───────────────────────────────────

def _flatten(record: dict, key_map: dict) -> dict:
    """Extract nested fields using dot-notation keys."""
    row = {}
    for dot_key, label in key_map.items():
        parts = dot_key.split(".")
        val = record
        for p in parts:
            if isinstance(val, dict):
                val = val.get(p)
            else:
                val = None
                break
        row[label] = val
    return row


def _to_df(records: list, col_map: dict) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=list(col_map.values()))
    return pd.DataFrame([_flatten(r, col_map) for r in records])


# ── Excel styling ─────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")  # dark navy
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
TOTAL_FILL    = PatternFill("solid", fgColor="E8F4FD")
TOTAL_FONT    = Font(bold=True, size=10)
BORDER_SIDE   = Side(style="thin", color="D0D0D0")
CELL_BORDER   = Border(bottom=BORDER_SIDE)


def _style_header(ws, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 45)


def _add_totals_row(ws, df: pd.DataFrame, numeric_cols: list[str]):
    """Append a totals row for numeric columns."""
    totals_row = ws.max_row + 1
    ws.cell(row=totals_row, column=1, value="TOTAL")
    ws.cell(row=totals_row, column=1).font = TOTAL_FONT

    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in numeric_cols:
            total = pd.to_numeric(df[col_name], errors="coerce").sum()
            cell = ws.cell(row=totals_row, column=col_idx, value=round(float(total), 2))
            cell.fill  = TOTAL_FILL
            cell.font  = TOTAL_FONT
            cell.number_format = '#,##0.00'


def _write_df(writer, sheet_name: str, df: pd.DataFrame,
              numeric_cols: list[str] | None = None):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _style_header(ws, len(df.columns))
    if numeric_cols:
        _add_totals_row(ws, df, numeric_cols)
    _auto_width(ws)


# ── Summary sheet ─────────────────────────────────────────────────────────

def _build_summary(data: dict) -> pd.DataFrame:
    def total(key, field):
        items = data.get(key) or []
        try:
            return round(sum(float(i.get(field, 0) or 0) for i in items), 2)
        except: return 0.0

    food_t = total("food_costs", "totalCostUsd")
    pkg_t  = total("packaging_costs", "totalCostUsd")
    toy_t  = total("toy_allocations", "totalCostUsd")
    mkt_t  = total("marketing_costs", "amountUsd")

    rows = [
        {"Category": "Food Costs",       "Entries": len(data.get("food_costs") or []),
         "Total Spend (USD)": food_t},
        {"Category": "Packaging Costs",  "Entries": len(data.get("packaging_costs") or []),
         "Total Spend (USD)": pkg_t},
        {"Category": "Toy Allocations",  "Entries": len(data.get("toy_allocations") or []),
         "Total Spend (USD)": toy_t},
        {"Category": "Marketing Costs",  "Entries": len(data.get("marketing_costs") or []),
         "Total Spend (USD)": mkt_t},
        {"Category": "GRAND TOTAL",      "Entries": "",
         "Total Spend (USD)": round(food_t + pkg_t + toy_t + mkt_t, 2)},
    ]
    return pd.DataFrame(rows)


# ── Main export function ───────────────────────────────────────────────────

def build_excel_report(data: dict) -> bytes:
    """
    Build the full multi-sheet Excel workbook from CostIQ API data.
    Returns raw bytes ready for Flask send_file().
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # 1. Summary
        _write_df(writer, "📊 Summary", _build_summary(data))

        # 2. Food Costs
        food_df = _to_df(data.get("food_costs") or [], FOOD_COLS)
        _write_df(writer, "🥩 Food Costs", food_df,
                  numeric_cols=["Quantity", "Unit Cost (USD)", "Total Cost (USD)"])

        # 3. Packaging Costs
        pkg_df = _to_df(data.get("packaging_costs") or [], PKG_COLS)
        _write_df(writer, "📦 Packaging Costs", pkg_df,
                  numeric_cols=["Quantity (units)", "Unit Cost (USD)", "Total Cost (USD)"])

        # 4. Toy Allocations
        toy_df = _to_df(data.get("toy_allocations") or [], TOY_COLS)
        _write_df(writer, "🎁 Toy Allocations", toy_df,
                  numeric_cols=["Qty", "Unit Cost (USD)", "Total Cost (USD)"])

        # 5. Marketing Costs
        mkt_df = _to_df(data.get("marketing_costs") or [], MKT_COLS)
        _write_df(writer, "📢 Marketing Costs", mkt_df,
                  numeric_cols=["Amount (USD)"])

        # 6. Campaigns
        camp_df = _to_df(data.get("campaigns") or [], CAMPAIGN_COLS)
        _write_df(writer, "🎯 Campaigns", camp_df,
                  numeric_cols=["Total Budget (USD)"])

        # 7. Suppliers
        supp_df = _to_df(data.get("suppliers") or [], SUPPLIER_COLS)
        _write_df(writer, "🚚 Suppliers", supp_df)

        # 8. Metadata sheet
        meta_df = pd.DataFrame([
            {"Field": "Generated At",    "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"Field": "Generated By",    "Value": "CostIQ AI ChatBot"},
            {"Field": "Data Source",     "Value": "CostIQ Spring Boot REST API"},
            {"Field": "Platform",        "Value": "GlobalBite Foods Inc. — CostIQ v1.0.0"},
            {"Field": "Auth",            "Value": "Keycloak 26 PKCE"},
        ])
        _write_df(writer, "ℹ️ Metadata", meta_df)

    output.seek(0)
    return output.read()